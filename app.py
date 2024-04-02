from flask import Flask, render_template, request, send_file
import oracledb
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime

app = Flask(__name__)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        CID_value = request.form.get('CID_value')

        # Establish connection to the Oracle database
        connection = oracledb.connect(
            user="mzadmin",
            password="VtoMz#016",
            dsn="10.137.167.11:1521/MZ"
        )

        # Define SQL query to fetch existing table data
        sql_query_fetch_data = """
          WITH dedup_data AS (
                SELECT 
                    b.CMPID, 
                    a.TESTCASEID, 
                    b.STEPID, 
                    a.DESCRIPTION AS A_DESCRIPTION, 
                    b.DESCRIPTION AS B_DESCRIPTION, 
                    b.VALUE1, 
                    b.VALUE2,
                    ROW_NUMBER() OVER (PARTITION BY a.TESTCASEID, b.STEPID ORDER BY a.TESTCASEID, b.STEPID) AS rn
                FROM 
                    AT_TESTCASEID a, 
                    AT_COMPARISIONMISMATCH b
                WHERE 
                    a.TESTCASEID = b.TESTCASEID 
                    AND b.CMPID = :1
            )
            SELECT 
                CMPID, 
                TESTCASEID, 
                STEPID, 
                A_DESCRIPTION, 
                B_DESCRIPTION, 
                VALUE1, 
                VALUE2
            FROM 
                dedup_data
            WHERE 
                rn = 1
        """

        cursor = connection.cursor()
        cursor.execute(sql_query_fetch_data, [CID_value])
        data = cursor.fetchall()

        workbook = Workbook()
        sheet1 = workbook.active
        sheet1.title = "Existing Data"
        sheet1.append(["CMPID", "TESTCASEID", "STEPID", "DESCRIPTION", "DESCRIPTION_1", "VALUE1", "VALUE2"])
        for row in data:
            sheet1.append(row)

        sql_query_pivot = """
         SELECT B_DESCRIPTION, 
                   COUNT(*) AS Count
            FROM (
                {}
            )
            GROUP BY B_DESCRIPTION
            ORDER BY Count Desc
        """.format(sql_query_fetch_data)

        cursor.execute(sql_query_pivot, [CID_value])
        pivot_data = cursor.fetchall()

        sheet2 = workbook.create_sheet(title="Pivot Table")
        sheet2.append(["TESTCASEID", "Count"])
        for row in pivot_data:
            sheet2.append(row)

        grand_total_row = [sum(row[1] for row in pivot_data)]
        sheet2.append(["GRAND TOTAL", grand_total_row[0]])

        blue_sky_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        for cell in sheet2[1]:
            cell.fill = blue_sky_fill

        for cell in sheet2[sheet2.max_row]:
            cell.fill = blue_sky_fill

        current_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"{CID_value}_report_on_{current_timestamp}.xlsx"
        workbook.save(filename=file_name)

        return send_file(file_name, as_attachment=True)

    return render_template('index.html')

if __name__ == '__main__':
    app.run(port=8034, debug=True)
